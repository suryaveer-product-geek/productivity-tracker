import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import os

EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "hourly_tracker.xlsx")
SHEET_NAME = "My Day Tracker"

# Time slots: 5:00 AM to 2:00 AM (21 slots)
TRACKED_HOURS = list(range(5, 24)) + [0, 1, 2]
TIME_LABELS = [f"{h:02d}:00" for h in TRACKED_HOURS]  # "05:00", "06:00" ... "02:00"

# Category colors
CAT_COLORS = {
    "Productive":   "A8D5B5",   # soft green
    "Work":         "A8C4E0",   # soft blue
    "Learning":     "C4B0E0",   # soft purple
    "Personal":     "F5D5A0",   # soft amber
    "Phone/Social": "F0A8A8",   # soft red
    "Rest":         "D5D5D5",   # soft gray
    "Sleep":        "B8C4CC",   # soft slate
    "Exercise":     "A8DDD5",   # soft teal
    "Meal":         "F5C8A0",   # soft orange
}

CAT_TEXT = {
    "Productive":   "1A5C30",
    "Work":         "1A3A5C",
    "Learning":     "3A1A5C",
    "Personal":     "5C3A00",
    "Phone/Social": "5C1A1A",
    "Rest":         "4A4A4A",
    "Sleep":        "2C3A42",
    "Exercise":     "0A3A35",
    "Meal":         "5C2E00",
}

# Column A is time labels, dates start at column 2
# Each date occupies 2 columns: [Category | Notes]
DATE_COL_WIDTH = 2  # category col + notes col

def _col_for_date(date_str, ws):
    """Find which column a date occupies, or add it if new."""
    # Scan row 1 for the date
    for col in range(2, ws.max_column + 1, DATE_COL_WIDTH):
        val = ws.cell(row=1, column=col).value
        if val == date_str:
            return col
    # Not found — add new date columns at the end
    new_col = max(2, ws.max_column + 1)
    # Snap to even column boundary
    if (new_col - 2) % DATE_COL_WIDTH != 0:
        new_col = 2 + (((new_col - 2) // DATE_COL_WIDTH) + 1) * DATE_COL_WIDTH
    _add_date_columns(ws, new_col, date_str)
    return new_col

def _add_date_columns(ws, start_col, date_str):
    """Add a new date header + empty category/notes columns."""
    cat_col  = start_col
    note_col = start_col + 1

    # Column widths
    ws.column_dimensions[get_column_letter(cat_col)].width  = 14
    ws.column_dimensions[get_column_letter(note_col)].width = 28

    # Date header spanning 2 cols (merge)
    ws.merge_cells(
        start_row=1, start_column=cat_col,
        end_row=1,   end_column=note_col
    )
    date_cell = ws.cell(row=1, column=cat_col, value=date_str)
    date_cell.fill      = PatternFill("solid", fgColor="C9A87C")
    date_cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    date_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sub-headers row 2
    cat_hdr  = ws.cell(row=2, column=cat_col,  value="Category")
    note_hdr = ws.cell(row=2, column=note_col, value="Notes")
    for cell in [cat_hdr, note_hdr]:
        cell.fill      = PatternFill("solid", fgColor="F5EFE6")
        cell.font      = Font(name="Calibri", bold=True, color="8B6A3E", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Pre-fill all time rows with Sleep default
    for i, label in enumerate(TIME_LABELS):
        row = i + 3   # data starts at row 3
        sleep_fill = PatternFill("solid", fgColor=CAT_COLORS["Sleep"])
        sleep_font = Font(name="Calibri", color=CAT_TEXT["Sleep"], size=10)
        center     = Alignment(horizontal="center", vertical="center")

        cat_cell  = ws.cell(row=row, column=cat_col,  value="Sleep")
        note_cell = ws.cell(row=row, column=note_col, value="")

        cat_cell.fill      = sleep_fill
        cat_cell.font      = sleep_font
        cat_cell.alignment = center

        note_cell.fill      = PatternFill("solid", fgColor="FAFAF8")
        note_cell.font      = Font(name="Calibri", color="AAAAAA", size=10, italic=True)
        note_cell.alignment = Alignment(vertical="center", wrap_text=True)

def _setup_sheet(ws):
    """Build the base sheet with time labels in column A."""
    # Freeze panes so row 1-2 and col A stay visible when scrolling
    ws.freeze_panes = "B3"

    # Column A — time labels
    ws.column_dimensions["A"].width = 10

    # Row heights
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 20

    # Column A header (rows 1+2 merged)
    ws.merge_cells("A1:A2")
    a1 = ws.cell(row=1, column=1, value="Time")
    a1.fill      = PatternFill("solid", fgColor="2C2A27")
    a1.font      = Font(name="Calibri", bold=True, color="C9A87C", size=11)
    a1.alignment = Alignment(horizontal="center", vertical="center")

    # Time label rows (start at row 3)
    for i, label in enumerate(TIME_LABELS):
        row  = i + 3
        cell = ws.cell(row=row, column=1, value=label)
        cell.fill      = PatternFill("solid", fgColor="F5F3EF")
        cell.font      = Font(name="Calibri", bold=True, color="4A4642", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 20

def init_excel():
    if not os.path.exists(EXCEL_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        _setup_sheet(ws)
        # Add today's date columns right away
        today_str = date.today().strftime("%d-%b")
        _add_date_columns(ws, 2, today_str)
        wb.save(EXCEL_PATH)
        print(f"[Tracker] Created Excel file: {EXCEL_PATH}")
    else:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(title=SHEET_NAME)
            _setup_sheet(ws)
        ws = wb[SHEET_NAME]
        # Ensure today's column exists
        today_str = date.today().strftime("%d-%b")
        _col_for_date(today_str, ws)
        wb.save(EXCEL_PATH)

def save_entry(slot, description, category, timestamp):
    try:
        wb   = openpyxl.load_workbook(EXCEL_PATH)
        ws   = wb[SHEET_NAME]

        today_str = timestamp.strftime("%d-%b")
        cat_col   = _col_for_date(today_str, ws)
        note_col  = cat_col + 1

        # Find the row for this time slot
        # slot format: "09:00 - 10:00" — match on start hour "09:00"
        start_label = slot.split(" - ")[0]   # e.g. "09:00"
        target_row  = None
        for i, label in enumerate(TIME_LABELS):
            if label == start_label:
                target_row = i + 3
                break

        if target_row is None:
            print(f"[Tracker] Could not find row for slot {slot}")
            return

        # Category cell
        color     = CAT_COLORS.get(category, "D5D5D5")
        txt_color = CAT_TEXT.get(category, "333333")
        cat_cell  = ws.cell(row=target_row, column=cat_col)
        cat_cell.value     = category
        cat_cell.fill      = PatternFill("solid", fgColor=color)
        cat_cell.font      = Font(name="Calibri", bold=True, color=txt_color, size=10)
        cat_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Notes cell
        note_cell = ws.cell(row=target_row, column=note_col)
        note_cell.value     = description if description else ""
        note_cell.font      = Font(name="Calibri", color="3A3632", size=10,
                                   italic=not bool(description))
        note_cell.fill      = PatternFill("solid", fgColor="FFFEFB")
        note_cell.alignment = Alignment(vertical="center", wrap_text=True)

        wb.save(EXCEL_PATH)
        print(f"[Tracker] Saved: {slot} → {category} | {description[:30] if description else '—'}")

    except Exception as e:
        print(f"[Tracker] Error saving entry: {e}")
